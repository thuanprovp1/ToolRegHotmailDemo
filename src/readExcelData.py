import pprint
import openpyxl
import os
import datetime
import pathlib

current_path = str(pathlib.Path().resolve())
path_of_excel = os.path.join(current_path, 'data.xlsx')
sheet_working = 'Sheet1'


def load_excel():
    return openpyxl.load_workbook(path_of_excel)


def read_info_in_excel():
    # list_info=['username','password','firstname','lastname','date of birth','row_hien_tai']
    list_all_info = []
    book = load_excel()
    # book.active = 1
    sheet = book[sheet_working]
    i = 2
    for i in range(i, sheet.max_row + 1):
        if sheet.cell(column=8, row=i).value is None:
            # if sheet.cell(column=8, row=i).value == 'success':
            get_date = datetime.datetime.strptime(str(sheet.cell(column=7, row=i).value),
                                                  "%Y-%m-%d %H:%M:%S")  # date of birth
            username = sheet.cell(column=1, row=i).value
            password = sheet.cell(column=2, row=i).value
            fullname = sheet.cell(column=3, row=i).value
            firstname = sheet.cell(column=4, row=i).value
            lastname = sheet.cell(column=5, row=i).value
            proxy = sheet.cell(column=6, row=i).value
            date_of_birth = get_date.strftime("%m/%d/%Y")
            current_row = i
            list_all_info.append(
                {'username': username, 'password': password, 'fullname': fullname,
                 'firstname': firstname, 'lastname': lastname, 'data_of_birth': date_of_birth, 'proxy': proxy,
                 'current_row': current_row})
    return list_all_info


def write_status(text, row_working, column_working):
    book = load_excel()
    sheet = book[sheet_working]
    sheet.cell(column=column_working, row=row_working).value = text
    save_excel(book)


def write_email(text, row_working):
    book = load_excel()
    sheet = book[sheet_working]
    sheet.cell(column=1, row=row_working).value = text
    save_excel(book)


def save_excel(book):
    book.save(path_of_excel)

# pprint.pprint(read_info_in_excel())
# write_status('test', 3, 8)
